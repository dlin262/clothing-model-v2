import os, io
from datetime import datetime
from flask import (Flask, render_template, request, jsonify,
                   send_file, session, redirect, url_for)
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'change-me-in-production-please')
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
ALLOWED_EXT = {'png', 'jpg', 'jpeg', 'gif', 'webp'}

OWNER_PIN = os.environ.get('OWNER_PIN', '1234')
STAFF_PIN = os.environ.get('STAFF_PIN', '5678')

DATABASE_URL = (os.environ.get('DATABASE_URL') or
                os.environ.get('RAILWAY_DATABASE_URL') or
                os.environ.get('POSTGRES_URL') or '')

def get_db():
    if DATABASE_URL:
        import psycopg2
        url = DATABASE_URL.replace('postgres://', 'postgresql://', 1)
        conn = psycopg2.connect(url)
        conn.autocommit = False
        return conn, 'pg'
    else:
        import sqlite3
        conn = sqlite3.connect('inventory.db')
        conn.row_factory = sqlite3.Row
        return conn, 'sqlite'

def fetchall(cur):
    rows = cur.fetchall()
    if not rows: return []
    if hasattr(rows[0], 'keys'): return [dict(r) for r in rows]
    cols = [d[0] for d in cur.description]
    return [dict(zip(cols, r)) for r in rows]

def fetchone(cur):
    row = cur.fetchone()
    if row is None: return None
    if hasattr(row, 'keys'): return dict(row)
    cols = [d[0] for d in cur.description]
    return dict(zip(cols, row))

def ph(db_type): return '%s' if db_type == 'pg' else '?'

def init_db():
    conn, db = get_db(); cur = conn.cursor()
    if db == 'pg':
        cur.execute('''CREATE TABLE IF NOT EXISTS items (
            id SERIAL PRIMARY KEY, name TEXT NOT NULL,
            quantity INTEGER NOT NULL DEFAULT 0,
            buy_price REAL NOT NULL DEFAULT 0, sale_price REAL NOT NULL DEFAULT 0,
            image_path TEXT, category TEXT DEFAULT 'General',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
        cur.execute('''CREATE TABLE IF NOT EXISTS sales (
            id SERIAL PRIMARY KEY, item_id INTEGER NOT NULL REFERENCES items(id) ON DELETE CASCADE,
            quantity_sold INTEGER NOT NULL, buy_price REAL NOT NULL DEFAULT 0,
            sale_price REAL NOT NULL, note TEXT,
            sold_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)''')
    else:
        cur.executescript('''
            CREATE TABLE IF NOT EXISTS items (
                id INTEGER PRIMARY KEY AUTOINCREMENT, name TEXT NOT NULL,
                quantity INTEGER NOT NULL DEFAULT 0, buy_price REAL NOT NULL DEFAULT 0,
                sale_price REAL NOT NULL DEFAULT 0, image_path TEXT,
                category TEXT DEFAULT "General", created_at TEXT DEFAULT CURRENT_TIMESTAMP);
            CREATE TABLE IF NOT EXISTS sales (
                id INTEGER PRIMARY KEY AUTOINCREMENT, item_id INTEGER NOT NULL,
                quantity_sold INTEGER NOT NULL, buy_price REAL NOT NULL DEFAULT 0,
                sale_price REAL NOT NULL, note TEXT, sold_at TEXT DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (item_id) REFERENCES items(id));''')
        for col in [('items','buy_price','REAL NOT NULL DEFAULT 0'),
                    ('sales','buy_price','REAL NOT NULL DEFAULT 0')]:
            try: cur.execute(f'ALTER TABLE {col[0]} ADD COLUMN {col[1]} {col[2]}')
            except: pass
    conn.commit(); conn.close()

def get_item_extras(conn, db, item_id):
    p = ph(db); cur = conn.cursor()
    cur.execute(f'SELECT COALESCE(SUM(quantity_sold),0) FROM sales WHERE item_id={p}', (item_id,))
    ts = float(cur.fetchone()[0] or 0)
    cur.execute(f'SELECT COALESCE(SUM(quantity_sold*sale_price),0) FROM sales WHERE item_id={p}', (item_id,))
    tr = float(cur.fetchone()[0] or 0)
    cur.execute(f'SELECT COALESCE(SUM(quantity_sold*buy_price),0) FROM sales WHERE item_id={p}', (item_id,))
    tc = float(cur.fetchone()[0] or 0)
    return ts, tr, tc

def require_login(roles=None):
    def decorator(fn):
        from functools import wraps
        @wraps(fn)
        def wrapper(*args, **kwargs):
            role = session.get('role')
            if not role: return redirect(url_for('login'))
            if roles and role not in roles: return jsonify({'error': 'Forbidden'}), 403
            return fn(*args, **kwargs)
        return wrapper
    return decorator

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        pin = request.form.get('pin', '')
        if pin == OWNER_PIN: session['role'] = 'owner'; return redirect(url_for('index'))
        elif pin == STAFF_PIN: session['role'] = 'staff'; return redirect(url_for('staff'))
        return render_template('login.html', error='Wrong PIN')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear(); return redirect(url_for('login'))

@app.route('/')
@require_login()
def index():
    if session.get('role') == 'staff': return redirect(url_for('staff'))
    return render_template('index.html')

@app.route('/staff')
@require_login()
def staff():
    return render_template('staff.html')

@app.route('/api/items', methods=['GET'])
@require_login()
def get_items():
    status = request.args.get('status', 'all')
    conn, db = get_db(); cur = conn.cursor()
    if status == 'instock': cur.execute('SELECT * FROM items WHERE quantity>0 ORDER BY name')
    elif status == 'outofstock': cur.execute('SELECT * FROM items WHERE quantity=0 ORDER BY name')
    else: cur.execute('SELECT * FROM items ORDER BY name')
    rows = fetchall(cur)
    items = []
    for item in rows:
        item['profit_per_unit'] = float(item['sale_price']) - float(item['buy_price'])
        ts, tr, tc = get_item_extras(conn, db, item['id'])
        item['total_sold'] = ts; item['total_revenue'] = tr; item['total_profit'] = tr - tc
        item['buy_price'] = float(item['buy_price']); item['sale_price'] = float(item['sale_price'])
        item['quantity'] = int(item['quantity'])
        items.append(item)
    conn.close(); return jsonify(items)

@app.route('/api/items', methods=['POST'])
@require_login()
def add_item():
    name = request.form.get('name', '').strip()
    if not name: return jsonify({'error': 'Name required'}), 400
    quantity = int(float(request.form.get('quantity', 0)))
    buy_price = float(request.form.get('buy_price', 0))
    sale_price = float(request.form.get('sale_price', 0))
    category = request.form.get('category', 'General').strip()
    image_path = None
    if 'image' in request.files:
        f = request.files['image']
        if f and f.filename and ('.' in f.filename and f.filename.rsplit('.', 1)[1].lower() in ALLOWED_EXT):
            fname = secure_filename(f"{datetime.now().strftime('%Y%m%d%H%M%S')}.jpg")
            os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
            save_path = os.path.join(app.config['UPLOAD_FOLDER'], fname)
            try:
                from PIL import Image, ExifTags
                img = Image.open(f)
                try:
                    exif = img._getexif()
                    if exif:
                        for tag, val in exif.items():
                            if ExifTags.TAGS.get(tag) == 'Orientation':
                                if val == 3:   img = img.rotate(180, expand=True)
                                elif val == 6: img = img.rotate(270, expand=True)
                                elif val == 8: img = img.rotate(90,  expand=True)
                except: pass
                img.thumbnail((800, 800), Image.LANCZOS)
                if img.mode in ('RGBA', 'P'): img = img.convert('RGB')
                img.save(save_path, 'JPEG', quality=75, optimize=True)
            except ImportError:
                f.seek(0); f.save(save_path)
            image_path = f"static/uploads/{fname}"
    conn, db = get_db(); cur = conn.cursor(); p = ph(db)
    cur.execute(f'INSERT INTO items (name,quantity,buy_price,sale_price,image_path,category) VALUES ({p},{p},{p},{p},{p},{p})',
                (name, quantity, buy_price, sale_price, image_path, category))
    conn.commit()
    if db == 'pg': cur.execute('SELECT lastval()'); new_id = cur.fetchone()[0]
    else: new_id = cur.lastrowid
    conn.close(); return jsonify({'success': True, 'id': new_id})

@app.route('/api/items/<int:item_id>', methods=['PUT'])
@require_login()
def update_item(item_id):
    data = request.get_json(); conn, db = get_db(); cur = conn.cursor(); p = ph(db)
    cur.execute(f'SELECT * FROM items WHERE id={p}', (item_id,))
    item = fetchone(cur)
    if not item: conn.close(); return jsonify({'error': 'Not found'}), 404
    cur.execute(f'UPDATE items SET name={p},quantity={p},buy_price={p},sale_price={p},category={p} WHERE id={p}',
                (data.get('name', item['name']), data.get('quantity', item['quantity']),
                 data.get('buy_price', item['buy_price']), data.get('sale_price', item['sale_price']),
                 data.get('category', item['category']), item_id))
    conn.commit(); conn.close(); return jsonify({'success': True})

@app.route('/api/items/<int:item_id>', methods=['DELETE'])
@require_login(roles=['owner'])
def delete_item(item_id):
    conn, db = get_db(); cur = conn.cursor(); p = ph(db)
    cur.execute(f'DELETE FROM sales WHERE item_id={p}', (item_id,))
    cur.execute(f'DELETE FROM items WHERE id={p}', (item_id,))
    conn.commit(); conn.close(); return jsonify({'success': True})

@app.route('/api/items/<int:item_id>/sell', methods=['POST'])
@require_login()
def record_sale(item_id):
    data = request.get_json()
    qty = int(data.get('quantity', 1)); price = float(data.get('sale_price', 0)); note = data.get('note', '')
    conn, db = get_db(); cur = conn.cursor(); p = ph(db)
    cur.execute(f'SELECT * FROM items WHERE id={p}', (item_id,))
    item = fetchone(cur)
    if not item: conn.close(); return jsonify({'error': 'Item not found'}), 404
    if int(item['quantity']) < qty: conn.close(); return jsonify({'error': 'Not enough stock'}), 400
    cur.execute(f'INSERT INTO sales (item_id,quantity_sold,buy_price,sale_price,note) VALUES ({p},{p},{p},{p},{p})',
                (item_id, qty, float(item['buy_price']), price, note))
    cur.execute(f'UPDATE items SET quantity=quantity-{p} WHERE id={p}', (qty, item_id))
    conn.commit(); conn.close(); return jsonify({'success': True})

@app.route('/api/stats')
@require_login()
def get_stats():
    conn, db = get_db(); cur = conn.cursor()
    cur.execute('SELECT COUNT(*) FROM items'); ti = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM items WHERE quantity>0'); ins = cur.fetchone()[0]
    cur.execute('SELECT COUNT(*) FROM items WHERE quantity=0'); oos = cur.fetchone()[0]
    cur.execute('SELECT COALESCE(SUM(quantity_sold*sale_price),0) FROM sales'); rev = float(cur.fetchone()[0] or 0)
    cur.execute('SELECT COALESCE(SUM(quantity_sold*buy_price),0) FROM sales'); cst = float(cur.fetchone()[0] or 0)
    cur.execute('SELECT COALESCE(SUM(quantity*sale_price),0) FROM items'); inv = float(cur.fetchone()[0] or 0)
    conn.close()
    return jsonify({'total_items':ti,'in_stock':ins,'out_of_stock':oos,
                    'total_revenue':rev,'total_profit':rev-cst,'inventory_value':inv})

@app.route('/api/sales')
@require_login()
def get_sales():
    conn, db = get_db(); cur = conn.cursor()
    cur.execute('''SELECT s.id, s.item_id, s.quantity_sold, s.buy_price, s.sale_price, s.note,
                   s.sold_at, i.name AS item_name, i.image_path,
                   (s.sale_price - s.buy_price) * s.quantity_sold AS profit
                   FROM sales s JOIN items i ON s.item_id=i.id ORDER BY s.sold_at DESC LIMIT 200''')
    rows = fetchall(cur); conn.close()
    for r in rows:
        r['profit'] = float(r['profit'] or 0)
        r['buy_price'] = float(r['buy_price']); r['sale_price'] = float(r['sale_price'])
        r['sold_at'] = str(r['sold_at'])
    return jsonify(rows)

@app.route('/api/report/excel')
@require_login(roles=['owner'])
def export_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({'error': 'openpyxl not installed'}), 500

    conn, db = get_db(); cur = conn.cursor()
    cur.execute('SELECT * FROM items ORDER BY name')
    all_items = fetchall(cur)
    cur.execute('''SELECT s.*, i.name AS item_name,
                   (s.sale_price-s.buy_price)*s.quantity_sold AS profit
                   FROM sales s JOIN items i ON s.item_id=i.id ORDER BY s.sold_at DESC''')
    all_sales = fetchall(cur)

    # Precompute extras for all items while connection is open
    extras = {}
    for item in all_items:
        extras[item['id']] = get_item_extras(conn, db, item['id'])
    conn.close()

    wb = openpyxl.Workbook()
    ACCENT='C8FF00'; BLACK='0A0A0A'; GREEN='22C55E'; RED='EF4444'; LGREY='F5F2ED'; MGREY='D4CFC8'

    def hfill(c): return PatternFill('solid', fgColor=c)
    def tborder():
        s = Side(style='thin', color=MGREY)
        return Border(left=s, right=s, top=s, bottom=s)

    def write_title(ws, title, cols):
        ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=cols)
        c = ws.cell(row=1,column=1,value=title)
        c.font=Font(bold=True,size=14,name='Calibri',color=BLACK); c.fill=hfill(ACCENT)
        c.alignment=Alignment(horizontal='left',vertical='center',indent=1); ws.row_dimensions[1].height=28
        ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=cols)
        sub=ws.cell(row=2,column=1,value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        sub.font=Font(size=9,color='888888',name='Calibri')

    def hdr_row(ws, row, headers):
        for i,h in enumerate(headers,1):
            c=ws.cell(row=row,column=i,value=h)
            c.fill=hfill(BLACK); c.font=Font(bold=True,color=ACCENT,name='Calibri',size=10)
            c.alignment=Alignment(horizontal='center',vertical='center')
        ws.row_dimensions[row].height=18

    def dcell(ws,row,col,val,fmt=None,color=None,alt=False):
        c=ws.cell(row=row,column=col,value=val)
        c.fill=hfill('F0EDE8' if alt else 'FFFFFF')
        c.font=Font(name='Calibri',size=10,color=color or BLACK)
        c.border=tborder(); c.alignment=Alignment(vertical='center')
        if fmt: c.number_format=fmt
        return c

    def setw(ws,widths):
        for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w

    def inv_sheet(ws, items, title):
        HDRS=['#','Name','Category','Buy Price (MMK)','Sale Price (MMK)',
              'Profit/Unit (MMK)','Qty','Status','Total Sold','Total Revenue (MMK)','Total Profit (MMK)']
        write_title(ws,title,len(HDRS)); hdr_row(ws,3,HDRS)
        for ri,item in enumerate(items,1):
            row=3+ri; alt=(ri%2==0)
            ts,tr,tc = extras.get(item['id'], (0,0,0))
            pu=float(item['sale_price'])-float(item['buy_price'])
            status='In Stock' if int(item['quantity'])>0 else 'Out of Stock'
            sc=GREEN if int(item['quantity'])>0 else RED
            vals=[ri,item['name'],item['category'],float(item['buy_price']),float(item['sale_price']),
                  pu,int(item['quantity']),status,ts,tr,tr-tc]
            fmts=[None,None,None,'#,##0','#,##0','#,##0',None,None,None,'#,##0','#,##0']
            for ci,(v,f) in enumerate(zip(vals,fmts),1):
                clr=sc if ci==8 else (GREEN if ci==11 and (tr-tc)>=0 else RED if ci==11 else None)
                dcell(ws,row,ci,v,fmt=f,color=clr,alt=alt)
        setw(ws,[4,30,14,16,18,18,7,13,11,20,18]); ws.freeze_panes='A4'

    ws1=wb.active; ws1.title='All Items'; inv_sheet(ws1,all_items,'MOCHI - All Inventory')
    ws2=wb.create_sheet('In Stock'); inv_sheet(ws2,[i for i in all_items if int(i['quantity'])>0],'MOCHI - In Stock')
    ws3=wb.create_sheet('Out of Stock'); inv_sheet(ws3,[i for i in all_items if int(i['quantity'])==0],'MOCHI - Out of Stock')

    ws4=wb.create_sheet('Sales History')
    SHDRS=['#','Item','Qty Sold','Buy (MMK)','Sale (MMK)','Revenue (MMK)','Profit (MMK)','Note','Date']
    write_title(ws4,'MOCHI - Sales History',len(SHDRS)); hdr_row(ws4,3,SHDRS)
    for ri,s in enumerate(all_sales,1):
        row=3+ri; alt=(ri%2==0)
        rev=float(s['quantity_sold'])*float(s['sale_price']); pft=float(s.get('profit') or 0)
        vals=[ri,s['item_name'],s['quantity_sold'],float(s['buy_price']),float(s['sale_price']),rev,pft,s.get('note') or '',str(s['sold_at'])[:16]]
        fmts=[None,None,None,'#,##0','#,##0','#,##0','#,##0',None,None]
        for ci,(v,f) in enumerate(zip(vals,fmts),1):
            clr=GREEN if ci==7 and pft>=0 else RED if ci==7 else None
            dcell(ws4,row,ci,v,fmt=f,color=clr,alt=alt)
    setw(ws4,[4,28,10,14,15,16,15,24,18]); ws4.freeze_panes='A4'

    ws5=wb.create_sheet('Summary'); write_title(ws5,'MOCHI - Business Summary',2)
    total_rev  = sum(float(s['quantity_sold'])*float(s['sale_price']) for s in all_sales)
    total_cost = sum(float(s['quantity_sold'])*float(s['buy_price'])  for s in all_sales)
    total_pft  = total_rev - total_cost
    margin     = (total_pft/total_rev*100) if total_rev else 0
    summary=[
        ('Total SKUs',len(all_items),''),
        ('In Stock',sum(1 for i in all_items if int(i['quantity'])>0),''),
        ('Out of Stock',sum(1 for i in all_items if int(i['quantity'])==0),''),
        ('','',''),
        ('Total Sales Transactions',len(all_sales),''),
        ('Total Revenue (MMK)',total_rev,'#,##0'),
        ('Total Cost (MMK)',total_cost,'#,##0'),
        ('Total Profit (MMK)',total_pft,'#,##0'),
        ('Profit Margin',margin/100,'0.00%'),
        ('','',''),
        ('Inventory Value at Sale Price (MMK)',sum(float(i['quantity'])*float(i['sale_price']) for i in all_items),'#,##0'),
        ('Inventory Value at Buy Price (MMK)', sum(float(i['quantity'])*float(i['buy_price'])  for i in all_items),'#,##0'),
    ]
    for ri,(label,value,fmt) in enumerate(summary,4):
        lc=ws5.cell(row=ri,column=1,value=label); vc=ws5.cell(row=ri,column=2,value=value if label else '')
        lc.font=Font(name='Calibri',size=11,bold=bool(label))
        pcolor=GREEN if label=='Total Profit (MMK)' and total_pft>=0 else RED if label=='Total Profit (MMK)' else BLACK
        vc.font=Font(name='Calibri',size=11,color=pcolor)
        if label: lc.fill=hfill(LGREY); vc.fill=hfill(LGREY)
        vc.alignment=Alignment(horizontal='right')
        if fmt: vc.number_format=fmt
    ws5.column_dimensions['A'].width=40; ws5.column_dimensions['B'].width=22

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    fname=f"mochi_report_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf,as_attachment=True,download_name=fname,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    os.makedirs('static/uploads', exist_ok=True)
    init_db()
    app.run(debug=True, port=5000)
